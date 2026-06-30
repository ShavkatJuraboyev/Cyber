from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from xml.sax.saxutils import escape as xml_escape
from utils.timezone import format_samarkand

def export_chats_to_txt(chats, filename="chats.txt"):
    with open(filename, "w", encoding="utf-8") as f:
        for row in chats:
            chat_id, title, chat_type, invite_link, is_admin = row[:5]
            bot_status = row[5] if len(row) > 5 else "unknown"
            f.write(f"ID: {chat_id} | {title} | {chat_type} | Admin: {bool(is_admin)} | Status: {bot_status} | link: {invite_link}\n")
    return filename


def export_chats_to_pdf(chats, filename="chats.pdf"):
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()
    story = []

    for row in chats:
        chat_id, title, chat_type, invite_link, is_admin = row[:5]
        bot_status = row[5] if len(row) > 5 else "unknown"
        text = f"ID: {chat_id} | {title} | {chat_type} | Admin: {bool(is_admin)} | Status: {bot_status} | link: {invite_link}"
        story.append(Paragraph(text, styles["Normal"]))

    doc.build(story)
    return filename


def _only_admin_referral_chats(chats):
    """Referral export uchun faqat bot admin bo‘lgan chatlarni qoldiradi."""
    return [row for row in chats if len(row) > 3 and int(row[3] or 0) == 1]


def export_referral_chats_to_txt(link_name, public_url, chats, filename="referral_chats.txt"):
    chats = _only_admin_referral_chats(chats)
    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"Giper ssilka: {link_name}\n")
        f.write(f"Ssilka: {public_url}\n")
        f.write(f"Bot admin bo‘lgan guruh/kanallar soni: {len(chats)}\n")
        f.write("=" * 60 + "\n\n")

        if not chats:
            f.write("Bu ssilka orqali bot admin qilingan guruh/kanal topilmadi.\n")
            return filename

        for i, row in enumerate(chats, start=1):
            chat_id, title, chat_type, is_admin, bot_status, added_at, added_by = row[:7]
            member_count = row[7] if len(row) > 7 else None
            f.write(f"{i}. {title or chat_id}\n")
            f.write(f"   ID: {chat_id}\n")
            f.write(f"   Turi: {chat_type}\n")
            f.write(f"   A’zolar soni: {member_count if member_count not in (None, '') else 'bazada yo‘q'}\n")
            f.write(f"   Status: {bot_status} | Admin: {bool(is_admin)}\n")
            f.write(f"   Qo‘shilgan: {format_samarkand(added_at)}\n")
            if added_by:
                f.write(f"   Kim qo‘shgan: {added_by}\n")
            f.write("\n")
    return filename


def export_referral_chats_to_pdf(link_name, public_url, chats, filename="referral_chats.pdf"):
    chats = _only_admin_referral_chats(chats)
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>Giper ssilka:</b> {xml_escape(str(link_name))}", styles["Title"]))
    story.append(Paragraph(f"<b>Ssilka:</b> {xml_escape(str(public_url))}", styles["Normal"]))
    story.append(Paragraph(f"<b>Bot admin bo‘lgan guruh/kanallar soni:</b> {len(chats)}", styles["Normal"]))
    story.append(Paragraph("<br/>", styles["Normal"]))

    if not chats:
        story.append(Paragraph("Bu ssilka orqali bot admin qilingan guruh/kanal topilmadi.", styles["Normal"]))
    else:
        for i, row in enumerate(chats, start=1):
            chat_id, title, chat_type, is_admin, bot_status, added_at, added_by = row[:7]
            member_count = row[7] if len(row) > 7 else None
            member_count_text = member_count if member_count not in (None, '') else 'bazada yo‘q'
            added_by_text = f" | Kim qo‘shgan: {added_by}" if added_by else ""
            text = (
                f"<b>{i}. {xml_escape(str(title or chat_id))}</b><br/>"
                f"ID: {chat_id} | Turi: {xml_escape(str(chat_type))} | A’zolar: {xml_escape(str(member_count_text))}<br/>"
                f"Status: {xml_escape(str(bot_status))} | Admin: {bool(is_admin)}<br/>"
                f"Qo‘shilgan: {xml_escape(format_samarkand(added_at))}{xml_escape(str(added_by_text))}"
            )
            story.append(Paragraph(text, styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

    doc.build(story)
    return filename


def export_users_to_txt(users, filename="users.txt"):
    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"Foydalanuvchilar soni: {len(users)}\n")
        f.write("=" * 60 + "\n\n")
        for i, row in enumerate(users, start=1):
            user_id, first_name, last_name, username, language_code, joined_at = row[:6]
            full_name = f"{first_name or ''} {last_name or ''}".strip() or "Noma’lum"
            username_text = f"@{username}" if username else "—"
            f.write(f"{i}. {full_name}\n")
            f.write(f"   ID: {user_id}\n")
            f.write(f"   Username: {username_text}\n")
            f.write(f"   Til: {language_code or '—'}\n")
            f.write(f"   Qo‘shilgan: {format_samarkand(joined_at)}\n\n")
    return filename


def export_users_to_pdf(users, filename="users.pdf"):
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("<b>Foydalanuvchilar ro‘yxati</b>", styles["Title"]))
    story.append(Paragraph(f"<b>Jami:</b> {len(users)}", styles["Normal"]))
    story.append(Paragraph("<br/>", styles["Normal"]))

    if not users:
        story.append(Paragraph("Foydalanuvchilar topilmadi.", styles["Normal"]))
    else:
        for i, row in enumerate(users, start=1):
            user_id, first_name, last_name, username, language_code, joined_at = row[:6]
            full_name = f"{first_name or ''} {last_name or ''}".strip() or "Noma’lum"
            username_text = f"@{username}" if username else "—"
            text = (
                f"<b>{i}. {xml_escape(str(full_name))}</b><br/>"
                f"ID: {user_id} | Username: {xml_escape(str(username_text))} | Til: {xml_escape(str(language_code or '—'))}<br/>"
                f"Qo‘shilgan: {xml_escape(format_samarkand(joined_at))}"
            )
            story.append(Paragraph(text, styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

    doc.build(story)
    return filename



def export_all_referral_chats_to_xlsx(referral_data, filename="referral_links.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Guruh kanallar"

    headers = [
        "TR",
        "Guruh/kanallar nomi",
        "Guruh/kanallar turi",
        "Id",
        "Azolar soni",
        "Kim qo'shgani",
        "Giper\nssilkasi nomi",
    ]

    blue = PatternFill("solid", fgColor="8DB4D9")
    title_blue = PatternFill("solid", fgColor="4F81BD")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Katta sarlavha
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Guruh/kanallar ro'yxati")
    title_cell.fill = title_blue
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Jadval shapka
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = blue
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 3
    tr = 1

    for item in referral_data:
        link_name = item.get("name") or "Nomsiz ssilka"
        chats = item.get("chats") or []

        for chat in chats:
            chat_id, title, chat_type, is_admin, bot_status, added_at, added_by = chat[:7]
            member_count = chat[7] if len(chat) > 7 else None

            values = [
                tr,
                title or "",
                chat_type or "",
                chat_id or "",
                member_count if member_count is not None else "",
                added_by or "",
                link_name,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row += 1
            tr += 1

    widths = [8, 28, 28, 18, 18, 22, 18]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 45
    ws.freeze_panes = "A3"

    wb.save(filename)
    return filename